"""Paper Trading Report Generator — weekly and monthly performance summaries.

Generates Excel (.xlsx) and CSV reports for paper-trading mode, including:

- Trade ledger (entry/exit, P&L per trade)
- Summary statistics (win rate, avg win/loss, max drawdown, Sharpe estimate)
- Equity curve data
- Sector breakdown

Usage::

    from src.reports.paper_report import PaperReportGenerator
    gen = PaperReportGenerator(db_manager)
    gen.generate_weekly_report()   # → data/reports/weekly_YYYY-WW.xlsx
    gen.generate_monthly_report()  # → data/reports/monthly_YYYY-MM.xlsx
"""
from __future__ import annotations

import csv
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, List, Optional
from zoneinfo import ZoneInfo

IST = ZoneInfo("Asia/Kolkata")
_log = logging.getLogger("reports.paper_report")

_REPORTS_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "reports"


class PaperReportGenerator:
    """Generates performance reports from paper-trading data stored in SQLite.

    Args:
        db_manager: :class:`~src.database.db_manager.DatabaseManager` instance.
        capital:    Starting capital used for return calculations (default ₹1,00,000).
    """

    def __init__(self, db_manager: Any, capital: float = 100_000) -> None:
        self.db = db_manager
        self._capital = capital
        _REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------ #
    # Public API                                                           #
    # ------------------------------------------------------------------ #

    def generate_weekly_report(self) -> dict:
        """Generate a weekly report for the current ISO week.

        Returns a summary dict and writes an Excel/CSV file to
        ``data/reports/weekly_YYYY-WW.xlsx``.
        """
        now    = datetime.now(IST)
        monday = now - timedelta(days=now.weekday())
        sunday = monday + timedelta(days=6)
        label  = now.strftime("W%W_%Y")

        trades = self._get_trades_in_range(
            monday.strftime("%Y-%m-%d"),
            sunday.strftime("%Y-%m-%d"),
        )

        summary = self._compute_summary(trades, label, "weekly")
        self._write_csv(trades, summary, f"weekly_{label}")
        self._write_excel(trades, summary, f"weekly_{label}")

        _log.info("Weekly report generated: %s (%d trades)", label, len(trades))
        return summary

    def generate_monthly_report(self) -> dict:
        """Generate a monthly report for the current month.

        Returns a summary dict and writes to ``data/reports/monthly_YYYY-MM.xlsx``.
        """
        now   = datetime.now(IST)
        label = now.strftime("%Y-%m")
        first = now.replace(day=1)
        # Last day of month: go to next month day-1
        if now.month == 12:
            last = now.replace(year=now.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last = now.replace(month=now.month + 1, day=1) - timedelta(days=1)

        trades = self._get_trades_in_range(
            first.strftime("%Y-%m-%d"),
            last.strftime("%Y-%m-%d"),
        )

        summary = self._compute_summary(trades, label, "monthly")
        self._write_csv(trades, summary, f"monthly_{label}")
        self._write_excel(trades, summary, f"monthly_{label}")

        _log.info("Monthly report generated: %s (%d trades)", label, len(trades))
        return summary

    # ------------------------------------------------------------------ #
    # Internal helpers                                                     #
    # ------------------------------------------------------------------ #

    def _get_trades_in_range(self, start_date: str, end_date: str) -> list[dict]:
        """Fetch trades from DB that were opened or closed in the given range."""
        try:
            # Get all trades, then filter by date range
            all_trades = self.db.get_trade_history(days=90)
            result = []
            for t in all_trades:
                entry = t.entry_date or t.created_at or ""
                if entry[:10] < start_date:
                    continue
                if entry[:10] > end_date:
                    continue
                result.append(t.to_dict())
            return result
        except Exception as exc:
            _log.error("Failed to fetch trades for report: %s", exc)
            return []

    def _compute_summary(
        self, trades: list[dict], label: str, period: str
    ) -> dict:
        """Compute performance summary statistics from a list of trade dicts."""
        closed = [t for t in trades if t.get("exit_date")]
        open_  = [t for t in trades if not t.get("exit_date")]

        pnls = [t.get("pnl", 0.0) or 0.0 for t in closed]
        wins = [p for p in pnls if p > 0]
        losses = [p for p in pnls if p <= 0]

        total_pnl     = sum(pnls)
        total_trades   = len(closed)
        winning_trades = len(wins)
        losing_trades  = len(losses)
        win_rate       = (winning_trades / total_trades * 100) if total_trades else 0.0
        avg_win        = (sum(wins) / len(wins)) if wins else 0.0
        avg_loss       = (sum(losses) / len(losses)) if losses else 0.0
        profit_factor  = (sum(wins) / abs(sum(losses))) if losses and sum(losses) != 0 else float("inf")

        # Max drawdown (sequential P&L series)
        max_dd = 0.0
        peak   = 0.0
        cumulative = 0.0
        for p in pnls:
            cumulative += p
            if cumulative > peak:
                peak = cumulative
            dd = peak - cumulative
            if dd > max_dd:
                max_dd = dd

        # Return %
        return_pct = (total_pnl / self._capital * 100) if self._capital else 0.0

        # Sector breakdown
        sector_pnl: dict[str, float] = {}
        for t in closed:
            # Extract sector from strategy_signal JSON if available
            sig = t.get("strategy_signal", "")
            sector = "UNKNOWN"
            if isinstance(sig, str) and "sector" in sig:
                try:
                    import json
                    parsed = json.loads(sig)
                    sector = parsed.get("sector", "UNKNOWN")
                except Exception:
                    pass
            sector_pnl[sector] = sector_pnl.get(sector, 0.0) + (t.get("pnl") or 0.0)

        return {
            "period":          period,
            "label":           label,
            "total_trades":    total_trades,
            "open_trades":     len(open_),
            "winning_trades":  winning_trades,
            "losing_trades":   losing_trades,
            "win_rate_pct":    round(win_rate, 2),
            "total_pnl":       round(total_pnl, 2),
            "return_pct":      round(return_pct, 2),
            "avg_win":         round(avg_win, 2),
            "avg_loss":        round(avg_loss, 2),
            "profit_factor":   round(profit_factor, 2) if profit_factor != float("inf") else "Inf",
            "max_drawdown":    round(max_dd, 2),
            "capital":         self._capital,
            "sector_pnl":      sector_pnl,
        }

    def _write_csv(self, trades: list[dict], summary: dict, filename: str) -> Path:
        """Write trade ledger and summary to a CSV file."""
        path = _REPORTS_DIR / f"{filename}.csv"
        try:
            with path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)

                # Summary section
                writer.writerow(["=== PERFORMANCE SUMMARY ==="])
                for key, val in summary.items():
                    if key != "sector_pnl":
                        writer.writerow([key, val])
                writer.writerow([])

                # Trade ledger
                if trades:
                    headers = [
                        "symbol", "trade_type", "quantity", "price",
                        "stop_loss", "target_price", "entry_date",
                        "exit_date", "exit_price", "pnl", "pnl_percentage",
                        "holding_days", "status",
                    ]
                    writer.writerow(headers)
                    for t in trades:
                        writer.writerow([t.get(h, "") for h in headers])

            _log.debug("CSV report written: %s", path)
            return path
        except Exception as exc:
            _log.error("Failed to write CSV report: %s", exc)
            return path

    def _write_excel(self, trades: list[dict], summary: dict, filename: str) -> Path:
        """Write trade ledger and summary to an Excel file using openpyxl.

        Falls back gracefully to CSV-only if openpyxl is not installed.
        """
        path = _REPORTS_DIR / f"{filename}.xlsx"
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            _log.info("openpyxl not installed — Excel report skipped (CSV available)")
            return path

        try:
            wb = Workbook()

            # ── Summary sheet ────────────────────────────────────────────
            ws = wb.active
            ws.title = "Summary"
            header_font = Font(bold=True, size=12)
            green_fill  = PatternFill(start_color="C6EFCE", fill_type="solid")
            red_fill    = PatternFill(start_color="FFC7CE", fill_type="solid")

            ws.append(["Paper Trading Report", summary.get("label", "")])
            ws["A1"].font = Font(bold=True, size=14)
            ws.append([])

            for key, val in summary.items():
                if key == "sector_pnl":
                    continue
                ws.append([key.replace("_", " ").title(), val])

            # Sector breakdown
            ws.append([])
            ws.append(["Sector P&L Breakdown"])
            for sector, pnl in summary.get("sector_pnl", {}).items():
                row = ws.max_row + 1
                ws.append([sector, f"{pnl:,.2f}"])
                fill = green_fill if pnl > 0 else red_fill
                ws.cell(row=row, column=2).fill = fill

            # ── Trade Ledger sheet ───────────────────────────────────────
            ws2 = wb.create_sheet("Trade Ledger")
            headers = [
                "Symbol", "Type", "Qty", "Entry Price", "Stop Loss",
                "Target", "Entry Date", "Exit Date", "Exit Price",
                "P&L (INR)", "P&L %", "Days Held", "Status",
            ]
            ws2.append(headers)
            for cell in ws2[1]:
                cell.font = header_font

            for t in trades:
                pnl = t.get("pnl") or 0.0
                row_data = [
                    t.get("symbol", ""),
                    t.get("trade_type", ""),
                    t.get("quantity", 0),
                    t.get("price", 0.0),
                    t.get("stop_loss", ""),
                    t.get("target_price", ""),
                    t.get("entry_date", ""),
                    t.get("exit_date", ""),
                    t.get("exit_price", ""),
                    pnl,
                    t.get("pnl_percentage", ""),
                    t.get("holding_days", ""),
                    t.get("status", ""),
                ]
                ws2.append(row_data)
                row_num = ws2.max_row
                fill = green_fill if pnl > 0 else (red_fill if pnl < 0 else None)
                if fill:
                    ws2.cell(row=row_num, column=10).fill = fill

            # Auto-adjust column widths
            for ws_sheet in [ws, ws2]:
                for col in ws_sheet.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws_sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

            wb.save(path)
            _log.info("Excel report written: %s", path)
            return path

        except Exception as exc:
            _log.error("Failed to write Excel report: %s", exc)
            return path
